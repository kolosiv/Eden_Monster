"""Export Manager for Eden Analytics Pro - CSV, Excel, PDF exports."""

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from io import BytesIO

from utils.logger import get_logger

logger = get_logger(__name__)

# Check for optional dependencies
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


class ExportManager:
    """Manages data export in various formats."""
    
    def __init__(self, output_dir: str = "exports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    def export_to_csv(self, data: List[Dict], filename: str,
                     columns: List[str] = None) -> Path:
        """Export data to CSV file.
        
        Args:
            data: List of dictionaries to export
            filename: Output filename (without extension)
            columns: Optional list of columns to include
        
        Returns:
            Path to exported file
        """
        if not data:
            raise ValueError("No data to export")
        
        filepath = self.output_dir / f"{filename}.csv"
        
        # Determine columns
        if columns is None:
            columns = list(data[0].keys())
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            
            for row in data:
                filtered_row = {k: v for k, v in row.items() if k in columns}
                writer.writerow(filtered_row)
        
        logger.info(f"Exported {len(data)} rows to {filepath}")
        return filepath
    
    def export_to_json(self, data: Any, filename: str, pretty: bool = True) -> Path:
        """Export data to JSON file.
        
        Args:
            data: Data to export (dict, list, or any JSON-serializable object)
            filename: Output filename (without extension)
            pretty: Whether to format JSON with indentation
        
        Returns:
            Path to exported file
        """
        filepath = self.output_dir / f"{filename}.json"
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2 if pretty else None, default=str)
        
        logger.info(f"Exported data to {filepath}")
        return filepath
    
    def export_to_excel(self, data: List[Dict], filename: str,
                       sheet_name: str = "Data",
                       include_charts: bool = True) -> Optional[Path]:
        """Export data to Excel file with formatting and optional charts.
        
        Args:
            data: List of dictionaries to export
            filename: Output filename (without extension)
            sheet_name: Name of the main sheet
            include_charts: Whether to include charts if applicable
        
        Returns:
            Path to exported file or None if Excel not available
        """
        if not EXCEL_AVAILABLE:
            logger.warning("Excel export not available. Install openpyxl.")
            return None
        
        if not data:
            raise ValueError("No data to export")
        
        filepath = self.output_dir / f"{filename}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6C63FF", end_color="6C63FF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        columns = list(data[0].keys())
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column.replace('_', ' ').title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        for row_idx, row in enumerate(data, 2):
            for col_idx, column in enumerate(columns, 1):
                value = row.get(column, "")
                
                # Format specific types
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M')
                elif isinstance(value, float):
                    value = round(value, 4)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                
                # Color profit/loss values
                if 'profit' in column.lower() and isinstance(value, (int, float)):
                    if value > 0:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value < 0:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Adjust column widths
        for col_idx, column in enumerate(columns, 1):
            max_length = max(
                len(str(column)),
                max(len(str(row.get(column, ""))) for row in data)
            )
            ws.column_dimensions[chr(64 + col_idx)].width = min(max_length + 2, 30)
        
        # Add chart if data has numeric columns suitable for charting
        if include_charts:
            numeric_cols = [col for col in columns if any(
                isinstance(row.get(col), (int, float)) for row in data
            )]
            
            if numeric_cols and len(data) > 1:
                # Add summary sheet
                ws_summary = wb.create_sheet("Summary")
                self._add_summary_to_excel(ws_summary, data, numeric_cols)
        
        wb.save(filepath)
        logger.info(f"Exported {len(data)} rows to {filepath}")
        return filepath
    
    def _add_summary_to_excel(self, ws, data: List[Dict], numeric_cols: List[str]):
        """Add summary statistics to Excel worksheet."""
        ws['A1'] = "Summary Statistics"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        for col in numeric_cols:
            values = [r.get(col) for r in data if isinstance(r.get(col), (int, float))]
            if values:
                ws.cell(row=row, column=1, value=col.replace('_', ' ').title())
                ws.cell(row=row, column=2, value="Count:")
                ws.cell(row=row, column=3, value=len(values))
                ws.cell(row=row, column=4, value="Sum:")
                ws.cell(row=row, column=5, value=round(sum(values), 2))
                ws.cell(row=row, column=6, value="Avg:")
                ws.cell(row=row, column=7, value=round(sum(values) / len(values), 2))
                row += 1
    
    def export_to_pdf(self, data: List[Dict], filename: str,
                     title: str = "Report",
                     include_summary: bool = True) -> Optional[Path]:
        """Export data to PDF report.
        
        Args:
            data: List of dictionaries to export
            filename: Output filename (without extension)
            title: Report title
            include_summary: Whether to include summary statistics
        
        Returns:
            Path to exported file or None if PDF not available
        """
        if not PDF_AVAILABLE:
            logger.warning("PDF export not available. Install reportlab.")
            return None
        
        if not data:
            raise ValueError("No data to export")
        
        filepath = self.output_dir / f"{filename}.pdf"
        
        doc = SimpleDocTemplate(
            str(filepath),
            pagesize=A4,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#6C63FF'),
            spaceAfter=20
        )
        elements.append(Paragraph(title, title_style))
        
        # Subtitle with date
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.gray,
            spaceAfter=20
        )
        elements.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            subtitle_style
        ))
        
        # Summary if requested
        if include_summary:
            elements.append(Paragraph("Summary", styles['Heading2']))
            summary_data = self._generate_summary(data)
            summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C63FF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F5F5')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#DDDDDD')),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))
        
        # Data table
        elements.append(Paragraph("Detailed Data", styles['Heading2']))
        
        columns = list(data[0].keys())
        table_data = [[col.replace('_', ' ').title() for col in columns]]
        
        for row in data[:100]:  # Limit to 100 rows for PDF
            table_row = []
            for col in columns:
                value = row.get(col, "")
                if isinstance(value, float):
                    value = f"{value:.2f}"
                elif isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d')
                table_row.append(str(value)[:20])  # Truncate long values
            table_data.append(table_row)
        
        # Calculate column widths
        col_width = (7 * inch) / len(columns)
        data_table = Table(table_data, colWidths=[col_width] * len(columns))
        
        data_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C63FF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F8F8')]),
        ]))
        
        elements.append(data_table)
        
        # Footer
        elements.append(Spacer(1, 30))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.gray,
            alignment=1  # Center
        )
        elements.append(Paragraph(
            "Eden Analytics Pro - Hockey Arbitrage Intelligence",
            footer_style
        ))
        
        doc.build(elements)
        logger.info(f"Exported {len(data)} rows to {filepath}")
        return filepath
    
    def _generate_summary(self, data: List[Dict]) -> List[List[str]]:
        """Generate summary statistics for PDF report."""
        summary = [["Metric", "Value"]]
        
        summary.append(["Total Records", str(len(data))])
        
        # Find numeric columns and calculate stats
        if data:
            for key in data[0].keys():
                values = [r.get(key) for r in data if isinstance(r.get(key), (int, float))]
                if values and len(values) >= 3:
                    total = sum(values)
                    avg = total / len(values)
                    summary.append([f"{key.replace('_', ' ').title()} (Total)", f"{total:.2f}"])
                    summary.append([f"{key.replace('_', ' ').title()} (Avg)", f"{avg:.2f}"])
        
        return summary
    
    def export_bet_history(self, bets: List[Dict], format: str = "csv") -> Optional[Path]:
        """Export betting history in specified format.
        
        Args:
            bets: List of bet records
            format: 'csv', 'excel', 'pdf', or 'json'
        
        Returns:
            Path to exported file
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"bet_history_{timestamp}"
        
        if format == "csv":
            return self.export_to_csv(bets, filename)
        elif format == "excel":
            return self.export_to_excel(bets, filename, sheet_name="Bet History")
        elif format == "pdf":
            return self.export_to_pdf(bets, filename, title="Betting History Report")
        elif format == "json":
            return self.export_to_json(bets, filename)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def export_performance_report(self, stats: Dict, history: List[Dict],
                                 format: str = "pdf") -> Optional[Path]:
        """Export comprehensive performance report.
        
        Args:
            stats: Overall statistics dictionary
            history: Bet history data
            format: Export format
        
        Returns:
            Path to exported file
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"performance_report_{timestamp}"
        
        # Combine stats and history for export
        report_data = {
            "generated_at": datetime.now().isoformat(),
            "statistics": stats,
            "bet_history": history
        }
        
        if format == "json":
            return self.export_to_json(report_data, filename)
        elif format == "pdf" and PDF_AVAILABLE:
            return self._generate_performance_pdf(stats, history, filename)
        else:
            return self.export_to_json(report_data, filename)
    
    def _generate_performance_pdf(self, stats: Dict, history: List[Dict],
                                 filename: str) -> Path:
        """Generate detailed performance PDF report."""
        filepath = self.output_dir / f"{filename}.pdf"
        
        doc = SimpleDocTemplate(str(filepath), pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Title'],
            textColor=colors.HexColor('#6C63FF'),
            spaceAfter=30
        )
        elements.append(Paragraph("Performance Report", title_style))
        elements.append(Paragraph(
            f"Eden Analytics Pro - {datetime.now().strftime('%B %d, %Y')}",
            styles['Normal']
        ))
        elements.append(Spacer(1, 30))
        
        # Key Metrics
        elements.append(Paragraph("Key Metrics", styles['Heading2']))
        
        metrics = [
            ["Metric", "Value"],
            ["Total Profit", f"${stats.get('total_profit', 0):,.2f}"],
            ["ROI", f"{stats.get('roi', 0):.1f}%"],
            ["Win Rate", f"{stats.get('win_rate', 0):.1f}%"],
            ["Total Bets", str(stats.get('total_bets', 0))],
            ["Wins", str(stats.get('wins', 0))],
            ["Losses", str(stats.get('losses', 0))],
            ["Holes", str(stats.get('holes', 0))]
        ]
        
        metrics_table = Table(metrics, colWidths=[3*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C63FF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.gray),
        ]))
        elements.append(metrics_table)
        elements.append(Spacer(1, 30))
        
        # Recent Bets
        if history:
            elements.append(Paragraph("Recent Bets", styles['Heading2']))
            
            bet_data = [["Match", "Stake", "Result", "Profit"]]
            for bet in history[:20]:
                match = f"{bet.get('home_team', 'N/A')} vs {bet.get('away_team', 'N/A')}"[:30]
                stake = f"${bet.get('total_stake', 0):.2f}"
                result = bet.get('result', 'Pending')
                profit = f"${bet.get('profit', 0):+.2f}"
                bet_data.append([match, stake, result, profit])
            
            bet_table = Table(bet_data, colWidths=[3*inch, 1*inch, 1*inch, 1*inch])
            bet_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6C63FF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ]))
            elements.append(bet_table)
        
        doc.build(elements)
        return filepath


__all__ = ['ExportManager', 'PANDAS_AVAILABLE', 'EXCEL_AVAILABLE', 'PDF_AVAILABLE']
